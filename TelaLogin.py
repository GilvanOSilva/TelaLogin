from tkinter import *
from openpyxl import load_workbook
from time import sleep


class UserSys(Tk):
    def __init__(self, *args, **kargs):
        Tk.__init__(self, *args, **kargs)
        screen = Frame(self)
        screen.pack()
        self.frames = {}
        for F in (Login, Registry, ResetPSW, ControlPanel, Update, Remove):
            page_name = F.__name__
            frame = F(parent=screen, controller=self)
            self.frames[page_name] = frame
            frame.grid(row=0, column=0, sticky='nsew')
        self.show_frame('Login')

    def show_frame(self, page_name):
        frame = self.frames[page_name]
        frame.tkraise()
        frame.label_message['text'] = ''


class Login(Frame):
    label_message = None

    def __init__(self, parent, controller):
        Frame.__init__(self, parent)
        self.controller = controller
        pad_size = 5
        font_size = 5
        username = StringVar()
        password = StringVar()
        label_user = Label(self, text='Username', padx=pad_size, pady=pad_size, justify='left', font=font_size)
        label_user.grid(column=0, row=0, sticky='W')
        entry_user = Entry(self, textvariable=username, font=font_size)
        entry_user.grid(column=1, row=0)
        label_password = Label(self, text='Senha', padx=pad_size, pady=pad_size, justify='left', font=font_size)
        label_password.grid(column=0, row=1, sticky='W')
        entry_password = Entry(self, show='*', textvariable=password, font=font_size)
        entry_password.grid(column=1, row=1)
        button_login = Button(self, text='Acessar', command=lambda: self.access(entry_user.get(), entry_password.get()), padx=pad_size, pady=pad_size, font=font_size)
        button_login.grid(column=1, row=2, sticky='E , W')
        button_new_pass = Button(self, text='Esqueci a senha', command=lambda: controller.show_frame('ResetPSW'), padx=pad_size, pady=pad_size, font=font_size)
        button_new_pass.grid(column=1, row=3, sticky='E , W')
        button_registry = Button(self, text='Cadastrar Novo Username', command=lambda: controller.show_frame('Registry'), padx=pad_size, pady=pad_size, font=font_size)
        button_registry.grid(column=1, row=4, sticky='E , W')
        button_quit = Button(self, text='Sair', command=quit, padx=pad_size, pady=pad_size, font=font_size)
        button_quit.grid(column=1, row=5, sticky='E , W')
        Login.label_message = Label(self, text='', padx=pad_size, pady=pad_size, justify='left', relief='sunken', font=font_size)
        Login.label_message.grid(column=1, row=6, sticky='EW')

    def access(self, username, password):
        user_pos = globals()
        file_xl = load_workbook(filename='Users.xlsx')
        users = file_xl.active
        pos_tab = 1
        while users['A' + str(pos_tab)].value is not None:
            if username == users['A' + str(pos_tab)].value:
                if password == users['B' + str(pos_tab)].value:
                    user_pos['POS'] = pos_tab
                    Update.password.set(users['B' + str(POS)].value)
                    Update.name.set(users['D' + str(POS)].value)
                    Update.age.set(users['E' + str(POS)].value)
                    ControlPanel.label_welcome['text'] = 'Bem Vindo ' + users['A' + str(pos_tab)].value
                    Remove.label_welcome['text'] = ControlPanel.label_welcome['text']
                    sleep(0.5)
                    UserSys.show_frame(self.controller, 'ControlPanel')
                    break
                else:
                    Login.label_message['text'] = 'Senha incorreta.'
                    break
            pos_tab += 1
        if users['A' + str(pos_tab)].value is None:
            Login.label_message['text'] = 'User não está cadastrado.'


class Registry(Frame):
    label_message = None

    def __init__(self, parent, controller):
        Frame.__init__(self, parent)
        self.controller = controller
        pad_size = 5
        font_size = 5
        username = StringVar()
        password = StringVar()
        email = StringVar()
        name = StringVar()
        age = StringVar()
        label_user = Label(self, text='Username', padx=pad_size, pady=pad_size, justify='left', font=font_size)
        label_user.grid(column=0, row=0, sticky='W')
        entry_user = Entry(self, textvariable=username, font=font_size)
        entry_user.grid(column=1, row=0)
        label_password = Label(self, text='Senha', padx=pad_size, pady=pad_size, justify='left', font=font_size)
        label_password.grid(column=0, row=1, sticky='W')
        entry_password = Entry(self, textvariable=password, font=font_size)
        entry_password.grid(column=1, row=1)
        label_email = Label(self, text='E-Mail', padx=pad_size, pady=pad_size, justify='left', font=font_size)
        label_email.grid(column=0, row=2, sticky='W')
        entry_email = Entry(self, textvariable=email, font=font_size)
        entry_email.grid(column=1, row=2)
        label_name = Label(self, text='Nome', padx=pad_size, pady=pad_size, justify='left', font=font_size)
        label_name.grid(column=0, row=3, sticky='W')
        entry_name = Entry(self, textvariable=name, font=font_size)
        entry_name.grid(column=1, row=3)
        label_age = Label(self, text='Idade', padx=pad_size, pady=pad_size, justify='left', font=font_size)
        label_age.grid(column=0, row=4, sticky='W')
        entry_age = Entry(self, textvariable=age, font=font_size)
        entry_age.grid(column=1, row=4)
        button_registry = Button(self, text='Cadastar', command=lambda: self.registry_user(entry_user.get(), entry_password.get(), entry_email.get(), entry_name.get(), entry_age.get()), padx=pad_size, pady=pad_size, font=font_size)
        button_registry.grid(column=1, row=5, sticky='N , S, W, E')
        button_return = Button(self, text='Voltar', command=lambda: controller.show_frame('Login'), padx=pad_size, pady=pad_size, font=font_size)
        button_return.grid(column=1, row=6, sticky='N , S, W, E')
        Registry.label_message = Label(self, text='', padx=pad_size, pady=pad_size, justify='left', relief='sunken', font=font_size)
        Registry.label_message.grid(column=1, row=7, sticky='EW')

    def registry_user(self, username, password, email, name, age):
        file_xl = load_workbook(filename='Users.xlsx')
        users = file_xl.active
        pos_tab = 1
        exist_user = False
        exist_email = False
        while users['A' + str(pos_tab)].value is not None:
            if users['A' + str(pos_tab)].value == username:
                exist_user = True
            if users['C' + str(pos_tab)].value == email:
                exist_email = True
            pos_tab += 1
        if exist_user is True:
            Registry.label_message['text'] = 'Este usuário já está cadastrado.'
        elif len(username) < 6 or ' ' in username:
            Registry.label_message['text'] = 'Usuário não atende o crítério.'
        elif exist_email is True:
            Registry.label_message['text'] = 'Este E-mail já está cadastrado.'
        elif '@' not in email or '.' not in email or ' ' in email:
            Registry.label_message['text'] = 'E-mail não atende o crítério.'
        else:
            pos_tab = 1
            while users['A' + str(pos_tab)].value is not None:
                pos_tab += 1
            users['A' + str(pos_tab)] = username
            users['B' + str(pos_tab)] = password
            users['C' + str(pos_tab)] = email
            users['D' + str(pos_tab)] = name
            users['E' + str(pos_tab)] = age
            file_xl.save(filename='Users.xlsx')
            Registry.label_message['text'] = 'Cadastro efetuado com sucesso.'


class ResetPSW(Frame):
    label_message = None

    def __init__(self, parent, controller):
        Frame.__init__(self, parent)
        self.controller = controller
        pad_size = 5
        font_size = 5
        username = StringVar()
        email = StringVar()
        label_message = Label(self, text='Digite os dados abaixo e confirme.', padx=pad_size, pady=pad_size, justify='left', font=font_size)
        label_message.grid(column=0, columnspan=2, row=1, sticky='N , S')
        label_user = Label(self, text='Username', padx=pad_size, pady=pad_size, justify='left', font=font_size)
        label_user.grid(column=0, row=2, sticky='W')
        entry_user = Entry(self, textvariable=username, font=font_size)
        entry_user.grid(column=1, row=2)
        label_email = Label(self, text='E-Mail', padx=pad_size, pady=pad_size, justify='left', font=font_size)
        label_email.grid(column=0, row=3, sticky='W')
        entry_email = Entry(self, textvariable=email, font=font_size)
        entry_email.grid(column=1, row=3)
        button_submit = Button(self, text='Confirmar', command=lambda: self.reset_user_psw(entry_user.get(), entry_email.get()), padx=pad_size, pady=pad_size, font=font_size)
        button_submit.grid(column=1, row=4, sticky='N , S, W, E')
        button_return = Button(self, text='Voltar', command=lambda: controller.show_frame('Login'), padx=pad_size, pady=pad_size, font=font_size)
        button_return.grid(column=1, row=5, sticky='N , S, W, E')
        ResetPSW.label_message = Label(self, text='', padx=pad_size, pady=pad_size, justify='left', relief='sunken', font=font_size)
        ResetPSW.label_message.grid(column=1, row=6, sticky='EW')

    def reset_user_psw(self, username, email):
        file_xl = load_workbook(filename='Users.xlsx')
        users = file_xl.active
        pos_tab = 1
        while users['A' + str(pos_tab)].value is not None:
            if users['A' + str(pos_tab)].value == username:
                if users['C' + str(pos_tab)].value == email:
                    users['B' + str(pos_tab)] = 'Senha12345'
                    file_xl.save(filename='Users.xlsx')
                    ResetPSW.label_message['text'] = 'Sua senha é, Senha12345, clique em voltar.'
                    break
                else:
                    ResetPSW.label_message['text'] = 'E-mail inválido.'
                    break
            pos_tab += 1
        if users['A' + str(pos_tab)].value is None:
            ResetPSW.label_message['text'] = 'User não foi encontrado.'


class ControlPanel(Frame):
    label_message = None
    label_welcome = None

    def __init__(self, parent, controller):
        Frame.__init__(self, parent)
        self.controller = controller
        pad_size = 5
        font_size = 5
        contagem_total = ControlPanel.total_users(self)
        ControlPanel.label_message = Label(self, text='')
        ControlPanel.label_welcome = Label(self, text='', padx=pad_size, pady=pad_size, justify='left', font=font_size)
        ControlPanel.label_welcome.grid(column=0, row=0, sticky='W')
        button_update = Button(self, text='Alterar dados da conta', command=lambda: controller.show_frame('Update'), padx=pad_size, pady=pad_size, font=font_size)
        button_update.grid(column=0, row=1, sticky='E , W')
        button_remove = Button(self, text='Remover cadastro do sistema', command=lambda: controller.show_frame('Remove'), padx=pad_size, pady=pad_size, font=font_size)
        button_remove.grid(column=0, row=2, sticky='E , W')
        button_return = Button(self, text='Encerrar Sessão', command=lambda: controller.show_frame('Login'), padx=pad_size, pady=pad_size, font=font_size)
        button_return.grid(column=0, row=4, sticky='N , S, W, E')
        label_total_user = Label(self, text='Total de users cadastrados: ' + str(contagem_total), padx=pad_size, pady=pad_size, justify='left', font=font_size)
        label_total_user.grid(column=0, row=5, sticky='EW')


    def total_users(self):
        file_xl = load_workbook(filename='Users.xlsx')
        users = file_xl.active
        pos_tab = 1
        total = 0
        while users['A' + str(pos_tab)].value is not None:
            total += 1
            pos_tab += 1
        return total


class Update(Frame):
    name = None
    age = None
    label_message = None
    password = None

    def __init__(self, parent, controller):
        Frame.__init__(self, parent)
        self.controller = controller
        pad_size = 5
        font_size = 5
        Update.password = StringVar()
        Update.name = StringVar()
        Update.age = StringVar()
        label_password = Label(self, text='Senha', padx=pad_size, pady=pad_size, justify='left', font=font_size)
        label_password.grid(column=0, row=1, sticky='W')
        entry_password = Entry(self, textvariable=Update.password, font=font_size)
        entry_password.grid(column=1, row=1)
        label_name = Label(self, text='Nome', padx=pad_size, pady=pad_size, justify='left', font=font_size)
        label_name.grid(column=0, row=3, sticky='W')
        entry_name = Entry(self, textvariable=Update.name, font=font_size)
        entry_name.grid(column=1, row=3)
        label_age = Label(self, text='Idade', padx=pad_size, pady=pad_size, justify='left', font=font_size)
        label_age.grid(column=0, row=4, sticky='W')
        entry_age = Entry(self, textvariable=Update.age, font=font_size)
        entry_age.grid(column=1, row=4)
        button_update = Button(self, text='Alterar', command=lambda: self.update_user(entry_password.get(), entry_name.get(), entry_age.get()), padx=pad_size, pady=pad_size, font=font_size)
        button_update.grid(column=1, row=5, sticky='N , S, W, E')
        button_return = Button(self, text='Voltar', command=lambda: controller.show_frame('ControlPanel'), padx=pad_size, pady=pad_size, font=font_size)
        button_return.grid(column=1, row=6, sticky='N , S, W, E')
        Update.label_message = Label(self, text='', padx=pad_size, pady=pad_size, justify='left', relief='sunken', font=font_size)
        Update.label_message.grid(column=1, row=7, sticky='EW')

    def update_user(self, password, name, age):
        global POS
        file_xl = load_workbook(filename='Users.xlsx')
        users = file_xl.active
        users['B' + str(POS)] = password
        users['D' + str(POS)] = name
        users['E' + str(POS)] = age
        file_xl.save(filename='Users.xlsx')
        Update.label_message['text'] = 'Dados alterados.'


class Remove(Frame):
    label_message = None
    label_welcome = None

    def __init__(self, parent, controller):
        Frame.__init__(self, parent)
        self.controller = controller
        pad_size = 5
        font_size = 5
        Remove.label_message = Label(self, text='')
        Remove.label_welcome = Label(self, text='', padx=pad_size, pady=pad_size, justify='left', font=font_size)
        Remove.label_welcome.grid(column=0, row=0, sticky='W')
        label_message = Label(self, text='AVISO: Esta ação remove o seu user do sistema!', padx=pad_size, pady=pad_size, justify='left', font=font_size)
        label_message.grid(column=0, row=1, sticky='N , S')
        button_remove = Button(self, text='Remover', command=lambda: self.remove_user(), padx=pad_size, pady=pad_size, font=font_size)
        button_remove.grid(column=0, row=2, sticky='N , S, W, E')
        button_return = Button(self, text='Voltar', command=lambda: controller.show_frame('ControlPanel'), padx=pad_size, pady=pad_size, font=font_size)
        button_return.grid(column=0, row=3, sticky='NSEW')

    def remove_user(self):
        global POS
        file_xl = load_workbook(filename='Users.xlsx')
        users = file_xl.active
        users['A' + str(POS)] = None
        users['B' + str(POS)] = None
        users['C' + str(POS)] = None
        users['D' + str(POS)] = None
        users['E' + str(POS)] = None
        file_xl.save(filename='Users.xlsx')
        UserSys.show_frame(self.controller, 'Login')


POS = None
prog = UserSys()
prog.title('UserSys')
prog.geometry('400x300')
prog.mainloop()
